from keras.models import load_model
from PIL import Image, ImageOps #Install pillow instead of PIL
import numpy as np
import xlsxwriter
import cv2
import argparse
import openpyxl
import os
from PIL import Image

# Create an empty array to store the image characters
file_names = []

folder_path = "images"
# Disable scientific notation for clarity
np.set_printoptions(suppress=True)

# Load the model
model = load_model('keras_Model.h5', compile=False)

# Load the labels
class_names = open('labels.txt', 'r').readlines()

model2 = load_model('keras_Model2.h5', compile=False)

# Load the labels
class_names2 = open('labels2.txt', 'r').readlines()

# Creat a class to store the image attributes

class ImageAttributes:
  def __init__(self, name, lowangle, phototype):
    self.name = name
    self.lowangle = lowangle
    self.phototype = phototype

# Iterate through all the files in the current directory
for file in os.listdir(folder_path):
  # Check if the file is an image
  if file.endswith(".jpg") or file.endswith(".png"):
    # Open the image file
    image = Image.open(file).convert('RGB')

    # Create the array of the right shape to feed into the keras model
    data = np.ndarray(shape=(1, 224, 224, 3), dtype=np.float32)

    #resize the image to a 224x224 with the same strategy as in TM2:
    #resizing the image to be at least 224x224 and then cropping from the center
    size = (224, 224)
    image = ImageOps.fit(image, size, Image.Resampling.LANCZOS)

    #turn the image into a numpy array
    image_array = np.asarray(image)

    # Normalize the image
    normalized_image_array = (image_array.astype(np.float32) / 127.0) - 1

    # Load the image into the array
    data[0] = normalized_image_array

    # run the inference
    prediction = model.predict(data)
    index = np.argmax(prediction)
    class_name = class_names[index]
    confidence_score = prediction[0][index]
    prediction2 = model2.predict(data)
    index2 = np.argmax(prediction2)
    class_name2 = class_names2[index2]
    confidence_score2 = prediction2[0][index2]

    print('Class2:', class_name2, end='')
    print('Confidence score:', confidence_score2)

    print('Class:', class_name, end='')
    print('Confidence score:', confidence_score)

    # Add the file name to the array
    file_names.append(ImageAttributes(file,class_name,class_name2))

# Print the array of file names
print(file_names)

workbook = openpyxl.Workbook()

# Get the active sheet
sheet = workbook.active

for i, obj in enumerate(file_names):
    print(obj.name)
    sheet.cell(row=1, column= 1, value= "Image Name")
    sheet.cell(row=1, column= 3, value= "Photo Type")
    sheet.cell(row=1, column= 4, value= "Camera Angle")

    sheet.cell(row=i+1, column= 1, value=obj.name)
    sheet.cell(row=i+1, column= 3, value=obj.phototype)
    sheet.cell(row=i+1, column= 4, value=obj.lowangle)

# Save the workbook to an Excel file
workbook.save("data.xlsx")

